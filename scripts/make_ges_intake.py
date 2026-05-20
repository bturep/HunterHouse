#!/usr/bin/env python3
"""
make_ges_intake.py — generate the GES (Eric Gesinger Collection) intake workbook.

Produces GES_intake.xlsx at the repo root: a fill-in spreadsheet for cataloguing
the ~70 new GES items (Richard Hunter furniture drawings + photographs of Eric
Gesinger's built furniture), structured so the filled sheet can be read straight
back by a batch Wikibase-creation script.

Design notes (why it's built this way):
- ONE "Catalogue" sheet, not split by type — easier to batch, per Brandon's
  request. The Type column distinguishes drawing vs photograph; type-specific
  columns (drawing type / scale) are simply left blank on photo rows.
- Constants that never change (source collection = GES/Q182, designer = Hunter)
  are pre-filled so they aren't re-keyed 70 times.
- IDs are pre-assigned HH-GES-0001.. in cataloguing order (current collection-
  prefix convention; the HH-A-* scheme in WIKIBASE.md is legacy). They are
  permanent — don't renumber.
- Phases (Wikibase P62) are NOT QIDs here. You type a plain piece name in
  "Furniture piece / set"; the batch script creates one phase item per distinct
  value and links rows to it. No QID lookups while cataloguing.
- Dropdown validation on the controlled-vocabulary columns keeps the data clean
  enough to batch without a hand-fix pass.
- A second "Reference" sheet documents every column, the column→Wikibase-property
  map, the controlled vocabularies (QIDs), and a worked example row.

Re-run anytime:  python3 scripts/make_ges_intake.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "GES_intake.xlsx"

N_DRAWINGS = 35   # Richard Hunter furniture drawings
N_PHOTOS = 35     # photographs of Gesinger's built furniture
N_SPARE = 12      # blank trailing rows for "I found more than expected"

# --- column definitions -----------------------------------------------------
# (header, width, wikibase mapping shown in the cell comment, instruction text)
COLS = [
    ("ID", 14, "P2 (HH archive ID)",
     "Pre-assigned, permanent, sequential. Do NOT renumber or reuse. "
     "If you have fewer items than rows, delete spare rows from the BOTTOM only. "
     "More items: keep counting HH-GES-00xx in the spare rows."),
    ("Type", 13, "P1 instance of  (drawing=Q88 / photograph=Q89)",
     "Pick from the dropdown. Drawing = a Hunter design drawing. "
     "Photograph = a photo of the finished furniture."),
    ("Title / Label", 34, "Wikibase item label",
     "Short human title, e.g. 'Trestle table — elevation' or "
     "'Walnut armchair, three-quarter view'. Becomes the item's name."),
    ("Furniture piece / set", 24, "P62 part of  (phase created from this value)",
     "Plain name of the furniture piece/commission, e.g. 'Trestle table', "
     "'Dining chair', 'Bookcase'. Use the SAME wording for every drawing & "
     "photo of the same piece — that's how they get grouped."),
    ("Date", 13, "P82 date created  (drawings) / P82 (photos)",
     "Year is enough: 1985. Or 1985-06, or 1985-06-12 if known. "
     "Use n.d. if undated."),
    ("Date precision", 14, "drives the /9 /10 /11 time precision",
     "year (default) / month / day / n.d. — match how complete the Date is."),
    ("Creator / photographer", 24, "P80 creator",
     "Drawings: pre-filled Richard Hunter. Photographs: type the photographer's "
     "full name here when known (left blank on purpose — the 'Mary —' surname). "
     "A Wikibase person item gets created from this value."),
    ("Designed by", 22, "P80/P4 — furniture designer",
     "Pre-filled Richard Hunter for all rows (he designed the furniture in both "
     "the drawings and the photographed pieces). Change only if a piece isn't his."),
    ("Subject / maker documented", 24, "documents-work / depicts",
     "Photographs: pre-filled Eric Gesinger (the maker whose finished work the "
     "photo documents). Drawings: leave blank."),
    ("Drawing type", 18, "P88 drawing type  (drawings only)",
     "Drawings only — pick from dropdown. Leave blank on photo rows. "
     "'full-size template' = 1:1 shop pattern."),
    ("Medium", 22, "P91 medium",
     "e.g. 'Pencil on vellum', 'Ink on trace', 'Gelatin silver print', "
     "'Colour photograph'. Free text."),
    ("Scale", 12, "P90 scale  (drawings)",
     "Drawings: 1:1, 1:10, full size, NTS, etc. Blank for photos."),
    ("Dimensions", 16, "(physical size — held for a future property)",
     "Sheet size for drawings / print size for photos, e.g. '420 × 297 mm'. "
     "Optional but useful."),
    ("Built status", 15, "P92 built status",
     "Was the piece actually made? built / partially built / unbuilt / "
     "prototype only / unknown. Photos almost always = built."),
    ("Notes (description)", 46, "P100 notes (prose shown in the record pane)",
     "One or two plain sentences. What it is, anything notable. No metadata "
     "formatting — this is the human-readable lede."),
    ("Physical location", 20, "P142 physical location",
     "Archival path/box, e.g. 'S0005, SS0001, FL0002'. Fill as you re-house."),
    ("Legacy ID", 14, "P97 legacy identifier",
     "Any number/label already written on the physical item. Optional."),
    ("Image filename (R2 base)", 26, "drives P95/P96 image URLs",
     "Base name only, no extension/tier, e.g. "
     "'HH-GES-0001_TrestleTable_1985'. The batch step adds "
     "_thumb/_prev/_large + the R2 path."),
    ("Status", 16, "(your workflow tracking — not written to Wikibase)",
     "to do / described / image ready / ready to ingest."),
    ("Notes to Brandon / Claude", 30, "(not written to Wikibase)",
     "Per-row questions or flags for the cataloguing/ingest pass."),
]

TYPE_DV = '"drawing,photograph"'
PREC_DV = '"year,month,day,n.d."'
DRAW_DV = ('"plan,elevation,section,detail,perspective,axonometric,'
           'full-size template,working drawing,sketch,other"')
BUILT_DV = '"built,partially built,unbuilt,prototype only,unknown"'
STATUS_DV = '"to do,described,image ready,ready to ingest"'

HUNTER = "Richard Hunter (Q201)"
GESINGER = "Eric Gesinger (Q209)"

# --- styling palette --------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1E1C1A")
HEAD_FONT = Font(color="F3F1EC", bold=True, size=11)
PREFILL = PatternFill("solid", fgColor="EBE8DF")   # constant / pre-filled cells
ENTER_FILL = PatternFill("solid", fgColor="FFFFFF")  # cells you fill in
SECTION_FONT = Font(bold=True, size=12, color="8B3A2A")
THIN = Side(style="thin", color="D6D2C8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def build():
    wb = Workbook()

    # ===================== Sheet 1: Catalogue ==============================
    ws = wb.active
    ws.title = "Catalogue"
    ws.sheet_properties.tabColor = "8B3A2A"

    # header row
    for i, (head, width, _wb, instr) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i, value=head)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
        c.border = BORDER
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        ws.cell(row=1, column=i).comment = Comment(
            f"{head}\n→ Wikibase: {_wb}\n\n{instr}", "GES intake")
    ws.row_dimensions[1].height = 46
    ws.freeze_panes = "C2"  # keep header + ID + Type visible while scrolling

    # data validations
    dvs = {
        2: DataValidation(type="list", formula1=TYPE_DV, allow_blank=True),
        6: DataValidation(type="list", formula1=PREC_DV, allow_blank=True),
        10: DataValidation(type="list", formula1=DRAW_DV, allow_blank=True),
        14: DataValidation(type="list", formula1=BUILT_DV, allow_blank=True),
        19: DataValidation(type="list", formula1=STATUS_DV, allow_blank=True),
    }
    for dv in dvs.values():
        ws.add_data_validation(dv)

    total = N_DRAWINGS + N_PHOTOS
    last_row = 1 + total + N_SPARE

    for n in range(total + N_SPARE):
        r = 2 + n
        seq = n + 1
        is_stub = n < total
        is_drawing = n < N_DRAWINGS

        # defaults per row
        vals = {c: "" for c in range(1, len(COLS) + 1)}
        if is_stub:
            vals[1] = f"HH-GES-{seq:04d}"            # ID
            vals[2] = "drawing" if is_drawing else "photograph"
            vals[6] = "year"                          # Date precision default
            vals[8] = HUNTER                          # Designed by (always)
            vals[19] = "to do"                        # Status
            if is_drawing:
                vals[7] = HUNTER                      # Creator = Hunter
            else:
                vals[9] = GESINGER                   # Subject/maker = Gesinger
                # creator/photographer deliberately blank (the 'Mary —')

        for cidx in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=cidx, value=vals[cidx] or None)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            # shade pre-filled constants vs. cells you enter
            if vals[cidx]:
                cell.fill = PREFILL
            else:
                cell.fill = ENTER_FILL
        ws.row_dimensions[r].height = 30

        # attach dropdowns to this row's controlled cells
        for cidx, dv in dvs.items():
            dv.add(ws.cell(row=r, column=cidx))

    # a faint marker where the spare rows begin
    spare_start = 2 + total
    ws.cell(row=spare_start, column=1).comment = Comment(
        "Spare rows start here. Continue HH-GES numbering if you have more "
        "than ~70 items; otherwise leave blank or delete from the bottom.",
        "GES intake")

    # ===================== Sheet 2: Reference ==============================
    rf = wb.create_sheet("Reference")
    rf.sheet_properties.tabColor = "4F7A6B"
    rf.column_dimensions["A"].width = 26
    rf.column_dimensions["B"].width = 30
    rf.column_dimensions["C"].width = 60

    rrow = [1]

    def line(a="", b="", c="", bold=False, section=False):
        r = rrow[0]
        for col, val in ((1, a), (2, b), (3, c)):
            cell = rf.cell(row=r, column=col, value=val or None)
            cell.alignment = WRAP_TOP
            if section:
                cell.font = SECTION_FONT
            elif bold:
                cell.font = Font(bold=True)
        rrow[0] += 1

    line("GES intake — Eric Gesinger Collection", section=True)
    line("", "", "Richard Hunter furniture drawings + photographs of the "
         "finished furniture made by Eric Gesinger.")
    line()
    line("HOW TO USE", section=True)
    line("1.", "", "Fill the Catalogue sheet, one row per physical item. "
         "Work top-down; leave a column blank if it doesn't apply.")
    line("2.", "", "Use the dropdowns where they exist (Type, Date precision, "
         "Drawing type, Built status, Status).")
    line("3.", "", "Re-use the EXACT same 'Furniture piece / set' wording for "
         "every drawing and photo of the same piece — that grouping becomes "
         "a Wikibase phase.")
    line("4.", "", "Pre-filled grey cells are constants — only change them if "
         "a row is genuinely different.")
    line("5.", "", "Save and send the file back. I read it and generate the "
         "Wikibase entries (phases + items) for you to confirm.")
    line()
    line("CONSTANTS APPLIED AUTOMATICALLY", section=True)
    line("Source collection", "P79 → Q182", "Eric Gesinger Collection (GES) "
         "— set on every item; no column needed.")
    line("ID scheme", "P2", "HH-GES-0001.. in cataloguing order, "
         "zero-padded, never reused.")
    line("Designed by", "Hunter (Q201)", "Pre-filled on all rows.")
    line("Photographer", "(blank)", "Left empty on photo rows on purpose — "
         "the 'Mary —' surname is still unknown. Person item created later.")
    line()
    line("COLUMN → WIKIBASE PROPERTY MAP", section=True)
    line("Column", "Property", "Notes", bold=True)
    for head, _w, wbmap, instr in COLS:
        line(head, wbmap, instr)
    line()
    line("ITEM TYPES (P1)", section=True)
    line("Q88", "drawing", "")
    line("Q89", "photograph", "")
    line()
    line("DRAWING TYPES (P88)", section=True)
    line("", "", "Pick the closest from the Drawing-type dropdown. New "
         "furniture-specific types (e.g. 'full-size template') get created "
         "as Wikibase items at ingest if they don't exist yet.")
    for q, lbl in [("Q98", "plan"), ("Q99", "elevation"), ("Q100", "section"),
                   ("Q101", "detail")]:
        line(q, lbl, "")
    line("(new)", "perspective / axonometric / full-size template / "
         "working drawing / sketch", "created at ingest if absent")
    line()
    line("BUILT STATUS (P92)", section=True)
    for q, lbl in [("Q51", "built"), ("Q52", "partially built"),
                   ("Q53", "unbuilt"), ("Q54", "prototype only / theoretical"),
                   ("Q56", "unknown")]:
        line(q, lbl, "")
    line()
    line("PEOPLE", section=True)
    line("Q201", "Richard Morrow Hunter", "furniture designer (these drawings)")
    line("Q209", "Eric Gesinger", "furniture maker — collection source / "
         "subject of the photographs")
    line("(new)", "Photographer 'Mary —'", "person item created from the "
         "Creator/photographer column once the name is known")
    line()
    line("DATE FORMAT", section=True)
    line("1985", "year", "→ precision /9")
    line("1985-06", "month", "→ precision /10")
    line("1985-06-12", "day", "→ precision /11")
    line("n.d.", "n.d.", "no date claim written")
    line()
    line("WORKED EXAMPLE (one drawing + its photo)", section=True)
    line("Drawing row", "",
         "ID HH-GES-0001 · Type drawing · Title 'Trestle table — elevation' · "
         "Piece 'Trestle table' · Date 1985 · precision year · Creator Richard "
         "Hunter (Q201) · Designed by Richard Hunter · Drawing type elevation · "
         "Medium 'Pencil on vellum' · Scale 1:10 · Built status built · "
         "Notes 'Design elevation for the walnut trestle dining table built by "
         "Eric Gesinger.' · Image HH-GES-0001_TrestleTable_1985")
    line("Photo row", "",
         "ID HH-GES-0036 · Type photograph · Title 'Trestle table, finished' · "
         "Piece 'Trestle table' · Date 1986 · Creator (photographer name when "
         "known) · Designed by Richard Hunter · Subject/maker Eric Gesinger "
         "(Q209) · Medium 'Gelatin silver print' · Built status built · "
         "Notes 'The completed trestle table photographed in Gesinger's "
         "workshop.' · Image HH-GES-0036_TrestleTable_1986")

    for rr in range(1, rrow[0]):
        rf.row_dimensions[rr].height = 28

    wb.save(OUT)
    return OUT, total, N_SPARE, last_row


if __name__ == "__main__":
    out, total, spare, last = build()
    print(f"Wrote {out}")
    print(f"  {total} stub rows (HH-GES-0001..HH-GES-{total:04d}) "
          f"+ {spare} spare blank rows")
    print(f"  Catalogue rows 2..{last};  Reference sheet attached")
